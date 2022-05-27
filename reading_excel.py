import  datetime  
import openpyxl


class ExtractNamesFromExcel():
    def __init__(self) -> None:        
        self.today = datetime.date.today()
        self.today = self.today.strftime("%m/%d")
        self.workbook = openpyxl.load_workbook('birthdaysheet.xlsx')
        self.sheet = self.workbook['Sheet1']

    def getBirthdayNames(self):
        birthday_people = []
        ctr = 2
           
        while self.sheet.cell(row = ctr,column=2).value != None:
            
            date =self.sheet.cell(row = ctr,column=2).value
            date = date.strftime("%m/%d")

            if self.today == date :
                
                birthday_people.append(self.sheet.cell(row = ctr,column = 1 ).value)
            ctr += 1    
        return birthday_people


        

# print(sheet['A2'].value)
# print(sheet['A3'].value)

